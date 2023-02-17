from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from .models import MyModel

class Command(BaseCommand):
    help = 'Import data from an Excel file'

    def add_arguments(self, parser):
        parser.add_argument('C:/Users/krishna/Documents/anprsheet', help='C:/Users/krishna/Documents/anprsheet')

    def handle(self, *args, **options):
        file_path = options['C:/Users/krishna/Documents/anprsheet']
        workbook = load_workbook(file_path)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            MyModel.objects.create('Id','Data','Time','Direction','Images','plateImages','plateResult')
        self.stdout.write(self.style.SUCCESS('Data imported successfully'))